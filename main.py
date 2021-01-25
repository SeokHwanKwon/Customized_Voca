# -*- coding: utf-8 -*-
import openpyxl
import random
import os
from operator import itemgetter
import datetime

# 데이터 읽어와 리스트에 넣기
def read_xlsx():
    # 엑셀파일 열기
    file = os.getcwd() + "/voca.xlsx"
    book = openpyxl.load_workbook(filename=file, read_only=False, data_only=False)

    # sheet1 불러오기
    sheet1 = book.worksheets[0]

    data = []
    for row in sheet1.rows:
        word = []
        for j in range(2): # 단어, 뜻
            word.append(row[j].value)
        if row[2].value is None:
            word.append(int(5)) # default 가중치
        else:
            word.append(int(row[2].value)) # 가중치 입력

        if row[3].value is None:
            word.append(datetime.datetime.now())
        else:
            word.append(row[3].value) # 날짜
        data.append(word)

    book.close()
    return data


# update 엑셀
def update_xlsx(data, max_rows):
    file = os.getcwd() + "/voca.xlsx"
    book = openpyxl.Workbook()

    for n in range(1, (len(data) + 1)):
        book.worksheets[0].cell(row=n, column=1).value = data[n - 1][0]
        book.worksheets[0].cell(row=n, column=2).value = data[n - 1][1]
        book.worksheets[0].cell(row=n, column=3).value = data[n - 1][2]
        book.worksheets[0].cell(row=n, column=4).value = data[n - 1][3] #datetime.datetime.strptime(str(data[n - 1][3]), '%Y-%m-%d').date()

    for n in range(len(data), len(data) - max_rows):
        book.worksheets[0].cell(row=n, column=1).value = ' '
        book.worksheets[0].cell(row=n, column=2).value = ' '
        book.worksheets[0].cell(row=n, column=3).value = ' '
        book.worksheets[0].cell(row=n, column=4).value = ' '

    book.save(file)


# 점수 표시
def print_score(score, count):
    print("\n\n%d개를 맞추었습니다." % (score))
    print("당신의 점수는", "%.2f" % ((score * 100) / count) + "점 입니다.")


# arr이라는 리스트, 전체 개수, 자식노드 중 큰거
# To heapify subtree rooted at index i.
# n is size of heap
def heapify(data, n, i):
    largest = i  # Initialize largest as root
    l = 2 * i + 1  # left = 2*i + 1
    r = 2 * i + 2  # right = 2*i + 2

    # See if left child of root exists and is
    # greater than root
    if l < n and data[i][2] > data[l][2]:
        largest = l

        # See if right child of root exists and is
    # greater than root
    if r < n and data[largest][2] > data[r][2]:
        largest = r

        # Change root, if needed
    if largest != i:
        data[i], data[largest] = data[largest], data[i]  # swap

        # Heapify the root.
        heapify(data, n, largest)


def heapSort(data):
    n = len(data)

    # Build a maxheap.
    # Since last parent will be at ((n//2)-1) we can start at that location.
    # n부터 -1_2_까지 -1_3_만큼 숫자간격을 좁혀서
    for i in range(n // 2 - 1, -1, -1):
        heapify(data, n, i)

    # One by one extract elements
    for i in range(n - 1, 0, -1):
        data[i], data[0] = data[0], data[i]  # swap
        heapify(data, i, 0)

    return data



if __name__ == '__main__':

    # 데이터 읽어와 리스트에 넣기
    data = read_xlsx()
    data = heapSort(data)

    # 점수
    score = 0

    # 몇 문제를 풀 것인지
    q_num = int(input("몇 문제를 푸시겠습니까? : "))
    #q_list = random.sample(range(0, len(data)), q_num)
    q_list = [x for x in range(0, q_num)]

    # 테스트
    count = 0
    max_rows = len(data)
    mis_list = []


    while(count < q_num):
        print("\n")
        count += 1

        i = q_list.pop()

        # 보기
        select = []
        select.append(data[i])
        # 틀린 보기 만들기

        a = random.sample(range(0,len(data)),4)

        if i in a:
            a.remove(i)
        select.append(data[a[0]])
        select.append(data[a[1]])
        select.append(data[a[2]])
        random.shuffle(select)


        print(data[i][0])
        print("[1] " + select[0][1])
        print("[2] " + select[1][1])
        print("[3] " + select[2][1])
        print("[4] " + select[3][1])
        print("<정답을 모르면 5를 누르시오>")
        print()
        answer = int(input("번호 : ")) - 1

        if answer == 4:
            print("모르시군요!\n정답 : " + data[i][1])
            data[i][2] += 1
            mis_list.append(data[i])
        else:
            # 정답인 경우
            if select[answer][1] == data[i][1]:
                print("정답입니다!")
                score += 1
                # 가중치 변경
                data[i][2] -= 1
                # 만약, 가중치가 0이되어 지워야한다면
                if data[i][2] < 1:
                    print(data.pop(i))
                    # 문제의 번호들을 shift한다.
                    for x in range(0, len(q_list)):
                        q_list[x] -= 1

            # 오답인 경우
            else:
                print("오답, 정답은 " + data[i][1])
                data[i][2] += 1
                mis_list.append(data[i])



    # 테스트 종료
    print_score(score, count)

    # data를 가중치가 높은 순으로 정렬
    data.sort(key=itemgetter(2), reverse=True)

    x = input("\n종료하고 싶으시면 Enter를 눌러주세요")

    print("\n\n[틀리신 문제들의 목록]")
    for x in mis_list:
        print(x)

    update_xlsx(data, max_rows)